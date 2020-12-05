#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Nov 19 15:01:08 2018

@author: wei
"""

# import tkinter
import tkinter as tk
import tkinter.font as tkfont

# import matplotlib
from matplotlib.figure import Figure
from matplotlib.lines import Line2D
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk

#import node
from node import Node
from ORC_plot import calc_SaturationofCurve, calc_StatusofORC
from ORC_plot import ProcessPlot
from ORC_sample import initNode, setAndCalcNode
# import os
import os
from openpyxl import Workbook, load_workbook
import datetime
# import config
import config as cfg
import agilent_load as agilent
from realtime_data import data
# import gc


class P_I_Diagram(tk.Frame):
    offset_x = 50
    offset_y = 30

    def __init__(self, master=None):
        tk.Frame.__init__(self, master=None)

        self.canvasID = {}

        ''' load img and create canvas '''
        self.photo = cfg.import_photo()
        self.canvas = tk.Canvas(master,
                                width=self.photo.width(),
                                height=self.photo.height(),
                                bg='white')
        self.canvas.pack(expand=1, fill=tk.BOTH)
        ''' put gif image on canvas
        pic's upper left corner (NW) on the canvas is at x=50 y=10
        '''
        self.canvas.create_image(0, 0, image=self.photo, anchor=tk.NW)

        ''''font'''
        self.fontprop = tkfont.Font(
            size=cfg.GUI["fontsize"])  # bitstream charter or courier 10 pitch

        ''''set label'''
        self.set_Labels()

        # set label of efficiency
        # fontTitle = tkfont.Font(
        #     family='courier 10 pitch', size=30, weight='bold')
        # name = "429_ORC\nEff"
        # unit = "%"
        # self.canvas.create_text(280, 320, text="{} {} {}".format(
        #     name, " "*14, unit), fill='green', font=fontTitle)
        # self.canvasID["{}_value".format(name)] = self.canvas.create_text(
        #     300, 350, text="None", fill='green', font=fontTitle)

    def create_text(self, x, y, t):
        return self.canvas.create_text(x, y, text=t, fill='blue', font=self.fontprop)

    def set_Labels(self):
        for k, v in cfg.LABEL.items():
            if v["posx"] == 0 or v["posy"] == 0:
                continue
            sensor_type = k.split("_")[-1]
            # default unit
            if "T" in sensor_type:
                unit = "C"
            elif "P" in sensor_type:
                unit = "B"
            elif sensor_type in ["Win", "Wout", "Qin", "Qout"]:
                unit = "kW"
            elif sensor_type in ["Eff", "Ein", "Eout"]:
                unit = "%"
            elif sensor_type in ["mDot"]:
                unit = "kg/s"
            else:
                unit = "?"

            if unit in ["C", "B"]:
                self.create_text(v["posx"], v["posy"],
                                 f"{sensor_type:<7}{'':^5}{unit:>5}")
                self.canvasID[f"{k}"] = self.create_text(
                    v["posx"], v["posy"], f"{'None':^6}")
            else:
                self.create_text(v["posx"], v["posy"],
                                 f"{sensor_type:<5}{' '*16:^5}{unit:>5}")
                self.canvasID[f"{k}"] = self.create_text(
                    v["posx"], v["posy"], f"{'None':^6}")

    def update_value(self, name, value, n=1):
        itemID = self.canvasID[name]
        self.canvas.itemconfigure(itemID, text=str(round(value, n)))

    def update(self):
        # print("update P&ID")
        for name, value in data.items():
            if isinstance(value, Node):
                self.update_value(f"{name}_T", data[name].t)
                self.update_value(f"{name}_P", data[name].p)
            else:
                self.update_value(name, value)


class ORC_Figure(tk.Frame):
    def __init__(self, master=None):
        tk.Frame.__init__(self, master=None)

        self.lines = {}

        self._fig = Figure(
            figsize=(cfg.FIG["width"], cfg.FIG["height"]), dpi=100)
        self.canvas = FigureCanvasTkAgg(self._fig, master)
        self.canvas._tkcanvas.pack(side=tk.TOP, fill=tk.BOTH, expand=1)

        self.ax = self._fig.add_subplot(111)

        self.set_title_label()
        self.set_window_boundary()
        self.openGrid()

        self.toolbar = NavigationToolbar2Tk(self.canvas, master)
        self.toolbar.update()
        self.plot_Saturation_Curve()

        self.set_line()

    def set_title_label(self):
        xAxis = "s"
        yAxis = "T"
        title = {"T": "T, °C", "s": "s, (kJ/kg)*K"}
        self.ax.set_title("%s-%s Diagram" % (yAxis, xAxis))
        self.ax.set_xlabel(title[xAxis])
        self.ax.set_ylabel(title[yAxis])

    def set_window_boundary(self):
        self.ax.set_ylim(10, 150)
        self.ax.set_xlim(0.9, 1.9)

    def openGrid(self):
        self.ax.grid()

    def add_line(self, line=None, **kw):
        if line is None:
            line = Line2D([], [], **kw)

        return self.ax.add_line(line)

    def plot_Saturation_Curve(self):
        saturation_curve = calc_SaturationofCurve()
        self.lines["saturation_curve_left"] = self.add_line(
            saturation_curve[0])
        self.lines["saturation_curve_right"] = self.add_line(
            saturation_curve[1])

    def set_line(self):
        for name, attr in cfg.LINE.items():
            if attr["type"] == "o":
                self.lines[f"{name}"] = self.add_line(
                    linestyle='None', color="k", marker=".")
            elif attr["type"] == "s":
                self.lines[f"{name}"] = self.add_line(lw=1.3, color="g")
            elif attr["type"] == "p":
                self.lines[f"{name}"] = self.add_line(lw=1.3, color="g")
            elif attr["type"] == "l":
                if name == "heat":
                    self.lines[f"{name}"] = self.add_line(lw=2.0, color="r")
                elif name == "cool":
                    self.lines[f"{name}"] = self.add_line(lw=2.0, color="b")
            else:
                print(f"{name} config error")

    def update_line(self, line_name, line_type, points):
        s = []
        T = []
        if line_type == "o":
            for point_name in points:
                s.append(data[f"{point_name}"].s)
                T.append(data[f"{point_name}"].t)
        elif line_type == "s":
            process = ProcessPlot(
                data[f"{points[0]}"], data[f"{points[1]}"], 'isos')
            process.test_iso_line()
            process.calc_iso()
            s = process.Isa
            T = process.Ita
        elif line_type == "p":
            process = ProcessPlot(
                data[f"{points[0]}"], data[f"{points[1]}"], 'isop')
            process.test_iso_line()
            process.calc_iso()
            s = process.Isa
            T = process.Ita
        elif line_type == "l":
            s_list = []
            for node in data.values():
                if isinstance(node, Node):
                    s_list.append(node.s)
            if line_name == "heat":
                s.append([max(s_list), min(s_list)])
            elif line_name == "cool":
                s.append([min(s_list), max(s_list)])
            else:
                print("f{line_name} config warning")
            T.append([data[f"{points[0]}_Ti"], data[f"{points[1]}_To"]])

        self.lines[f"{line_name}"].set_data(s, T)

    def update(self):
        # print("update T-s Diagram")
        for name, attr in cfg.LINE.items():
            self.update_line(name, attr["type"], attr["point"])

        self.canvas.draw()


class Scan_button(tk.Frame):
    def __init__(self, master=None, *callbacks):
        tk.Frame.__init__(self, master=None)
        # gc.disable()
        self.update_funcs = []
        for func in callbacks:
            if callable(func):
                self.update_funcs.append(func)
            else:
                def func(): return print("this is not a func")
                self.update_funcs.append(func)

        self.is_click = False

        def btn_cmd_loop(func):
            if self.is_click:
                self.is_click = False
                varScan.set('stop2scan')
            else:
                self.is_click = True
                varScan.set('start2scan')
                func()

        varScan = tk.StringVar()
        varScan.set('stop2scan')

        labelScan = tk.Label(
            master,
            textvariable=varScan,
            bg='white',
            font=('Arial', 12),
            width=15, height=2)
        labelScan.pack()

        button = tk.Button(
            master,
            text='click me',
            width=15, height=2,
            command=lambda: btn_cmd_loop(self.update_diagram))
        button.pack()

        ''' init v34970A '''
        self.dev = agilent.test_V34972A()
        self.file = csv_file()
        # self.dev = agilent.V34972A()

    def call_update_funcs(self):
        for func in self.update_funcs:
            func()
        return 0

    def update_diagram(self, count=0):
        if self.is_click:
            print("----" * 5)
            # print(f"update_diagram {count}")
            self.dev.scan()
            self.calc_nodes()
            self.file.print_data()
            ''' update functions
            update P&ID
            update T-s diagram
            '''
            self.call_update_funcs()
            # save_data()
            # print(gc.get_count())
            self.after(300, self.update_diagram, count+1)

    def calc_nodes(self):
        # print("calc nodes and works")
        ''' calc nodes '''
        for name, value in data.items():
            if isinstance(value, Node):
                # print(name, data[name].p, data[name].t)
                data[name].pt()
        ''' calc WORK '''
        for name, node in cfg.FM.items():
            item0 = data[f"{node[0]}"]
            item1 = data[f"{node[1]}"]
            mDot = data["mDot"]
            data[f"{name}"] = (item1.h - item0.h) * mDot

        ''' calc efficiency '''
        data["Eff"] = ((data["Wout"] - data["Win"]) / data["Qin"]) * 100
        
        
class csv_file:
    def __init__(self):
        self.rowdata = [ data["pump-in"].t, data["pump-in"].p, data["evaporator-in"].t ]
        pass
    def print_data(self):
        print(self.rowdata)
        print(data["pump-in"].t, data["pump-in"].p, data["evaporator-in"].t)
    


def must_be_remove_save_data():
    # pwd = os.getcwd()
    pwd = "/home/wei/app/GUI-of-Organic-Rankine-Cycle-experiment"
    path = f"{pwd}/weiGUIData"
    # print(path)
    filename = f'{datetime.date.today()}.xlsx'
    # mk_exclusivefile(path, filename)
    workBook, workSheet = mk_exclusivefile(path, filename)
    
    lastCell = workSheet.cell(workSheet.max_row, 1).value
    global i
    # if lastCell != 'scan':
    #     i = lastCell
    # else:
    #     i = 0

    i = i + 1
    workSheet.append(["0"] * 20)
    workSheet.append(["0"] * 20)
    workSheet.append(["0"] * 20)
    workSheet.append(["0"] * 20)
    workSheet.append(["0"] * 20)

    workBook.save("{}".format(filename))
    
    print(i, id(workBook), workBook)


def save_data():
    pass

def transfer_file():
    ''' transfer lock file to experiment file
    When experiment is done,
    system will copy lock file into experiment file.
    '''
    pass

def create_csv_file_header():
    pass

def _mk_lock_file():
    ''' avoid users crash file
    Lock file just ORC GUI can used.
    Prevent users from crashing the system due to file modification.
    '''
    pass

def mk_exclusivefile(path, filename):
    ''' 創見專屬資料夾

    說明: 在路徑 path 中, 創見名為 dirname 的資料夾
    -----
    ex:
        path = '/home/wei/data/python/photo'
        dirname = 'good'
        os.path.isdir(path + '/' + dirname) # False
        mk_exclusivedir(path, dirname)
        os.path.isdir(path + '/' + dirname) # True
    '''

    if not os.path.isdir(path):
        os.mkdir(path)
    os.chdir(path)
    if not os.path.isfile(filename):
        workBook = Workbook()
        workSheet = workBook.active
        workSheet['a1'] = '實驗名稱'
        workSheet['a2'] = '實驗日期'
        workSheet['b2'] = datetime.date.today()
        workSheet['a3'] = '實驗說明(描述)'
        workSheet.append(['scan', 'time(real)',
                          '壓差', 'inlet(P)', 'inlet(T)', '密度', '次冷', 'h1',
                          'outlet(P)', 'outlet(T)', '飽和溫度', 'h2',
                          '壓差', 'inlet(P)', 'inlet(T)', '飽和溫度', '過熱', 'h3',
                          'outlet(P)', 'outlet(T)', '飽和溫度', 'h4',
                          'inlet(T)', 'outlet(T)', '高溫壓迫',
                          'inlet(T)', 'outlet(T)', '低溫壓迫',
                          'ORC效率(%)', 'mdot(kg/s)', 'time(s)', '聚集', 'operate'])
        workBook.save("{}".format(filename))

    else:
        workBook = load_workbook('{}'.format(filename))
        workSheet = workBook.active

    return workBook, workSheet


#    if not os.path.isdir(dirname):
#        os.mkdir('{}'.format(dirname))
if __name__ == "__main__":
    pass
    # for k in range(1000000):
    #     save_data()