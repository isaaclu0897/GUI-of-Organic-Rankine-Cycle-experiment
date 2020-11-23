#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Mon Jul  9 00:24:49 2018

@author: wei
"""

import visa
from openpyxl import Workbook
import datetime
from threading import Timer
from GUIObj import mk_exclusivefile


def scan_data(data, SM_dia, TH_dia):
    # =============================================================================
    # load the data
    # =============================================================================
    probe_type_TEMP, type_TEMP, ch_TEMP = 'TCouple', 'T', '@201:210'
    range_PRESS, resolution_PRESS, ch_PRESS = 10, 5.5, '@301:306'
    gain_PRESS, state_PRESS = 2.1, 1
#    offset_PRESS, label_PRESS = 0, 'BAR'

    rm = visa.ResourceManager()
    v34972A = rm.open_resource('USB0::0x0957::0x2007::MY49017447::0::INSTR')
#        idn_string = v34972A.query('*IDN?')

    # data = SendData()

    path = r'C:\Users\lab\Desktop\weiGUIData'
    filename = '{}.xlsx'.format(datetime.date.today())
#    global workBook, workSheet
    workBook, workSheet = mk_exclusivefile(path, filename)

    lastCell = workSheet.cell(workSheet.max_row, 1).value
    global i
    if lastCell != 'scan':
        i = lastCell
    else:
        i = 0

    def innerfunc(SM_dia, TH_dia):
        # scan temperature
        scans_TEMP = v34972A.query(':MEASure:TEMPerature? %s,%s,(%s)' % (
            probe_type_TEMP, type_TEMP, ch_TEMP))

        # scan pressure
        v34972A.write(':CONFigure:VOLTage:DC %G,%G,(%s)' %
                      (range_PRESS, resolution_PRESS, ch_PRESS))
        v34972A.write(':CALCulate:SCALe:GAIN %G,(%s)' % (gain_PRESS, ch_PRESS))
        v34972A.write(':CALCulate:SCALe:STATe %d,(%s)' %
                      (state_PRESS, ch_PRESS))
        scans_PRESS = v34972A.query(':READ?')

        # convert str to float
        readings_TEMP = [float(x) for x in scans_TEMP.split(',')]
        readings_PRESS = [float(x) for x in scans_PRESS.split(',')]

#        print(readings_TEMP, readings_PRESS)

        value = data.send(readings_TEMP, readings_PRESS)
        global i
        i += 1
        prefixList = [i, datetime.datetime.now().strftime("%H:%M:%S")]
        postfixList = [i * 3]

        workSheet.append(prefixList+value+postfixList)

        workBook.save("{}".format(filename))
        data.update(SM_dia, TH_dia)

    timer(innerfunc, 3, SM_dia, TH_dia)


def test_scan_data(data, SM_dia, TH_dia):

    #    data = SendData()

    # workBook = Workbook()
    # workSheet = workBook.active
    # workSheet['a1'] = '實驗名稱'
    # workSheet['a2'] = '實驗日期'
    # workSheet['b2'] = datetime.date.today()
    # workSheet['a3'] = '實驗說明(描述)'
    # workBook.save("./DillWithData/sample.xlsx")

    def innerfunc(data, SM_dia, TH_dia):
        readings_PRESS = [1.8, 9, 8.3, 2.3, 1.9, 2]
        readings_TEMP = [22, 25, 97, 64, 24, 68, 99, 89, 22, 24]

        value = data.send(readings_TEMP, readings_PRESS)
        print(value)
        data.update(SM_dia, TH_dia)

    innerfunc(data, SM_dia, TH_dia)


def timer(func, second=2, *arg):
    func(*arg)
    t = Timer(second, timer, args=(func, 3, *arg))
    t.setDaemon(True)

    if t.daemon and on_click_loop:
        t.start()
    else:
        #        del readings_TEMP, readings_PRESS
        return 0

def update_diagram():
    # update P_and_I_Diagram
    # update ORC_Figure
    pass
    
    def innerfunc(text):
        print(text)
        # readings_PRESS = [1.8, 9, 8.3, 2.3, 1.9, 2]
        # readings_TEMP = [22, 25, 97, 64, 24, 68, 99, 89, 22, 24]
    
        # value = data.send(readings_TEMP, readings_PRESS)
        # print(value)
        # data.update(SM_dia, TH_dia)

    # innerfunc()
    timer(innerfunc, 3, "good")


if __name__ == '__main__':
    import tkinter as tk
    from GUIObj import ORC_Figure, P_and_I_Diagram, Scan_button
    window = tk.Tk()
    window.title("Lab429, ORC for 500W, author:wei")
    tk.Label(window, text='this is ORC_GUI').pack()

    frame = tk.Frame(window).pack()

    # left and right frame
    frm_right = tk.Frame(frame)
    frm_right.pack(side='right')
    tk.Label(frm_right, text='frame right').pack()

    frm_left = tk.Frame(frame)
    frm_left.pack(side='left')
    tk.Label(frm_left, text='frame left').pack()

    # top and bottom of right frame
    frm_right_top = tk.Frame(frm_right)
    frm_right_top.pack(side='top')
    tk.Label(frm_right_top, text='frame right top').pack()
    frm_right_bottom = tk.Frame(frm_right)
    frm_right_bottom.pack(side='bottom')
    tk.Label(frm_right_bottom, text='frame right bottom').pack()

    SM_dia = P_and_I_Diagram(frm_left)
    # TH_dia = ORC_Figure(frm_right_top)
    # data = SendData()
    # data = SendData()

    frm_right_bottom_left = tk.Frame(frm_right_bottom)
    frm_right_bottom_left.pack(side='left')
    tk.Label(frm_right_bottom_left, text='frame right bottom left').pack()
    
    frm_right_bottom_right = tk.Frame(frm_right_bottom)
    frm_right_bottom_right.pack(side='right')
    tk.Label(frm_right_bottom_right, text='frame right bottom right').pack()

    
    Scan_button(frm_right_bottom_left)
    
    # import sys

    # def close(event):
    #     window.withdraw() # if you want to bring it back
    #     sys.exit() # if you want to exit the entire thing

    # window.bind('<Escape>', close)
    # varScan = tk.StringVar()
    # labelScan = tk.Label(frm_right_bottom_left, textvariable=varScan, bg='white',
    #                      font=('Arial', 12), width=15, height=2)
    # labelScan.pack()

    # varmdotWater = tk.StringVar()
    # labelmdotWater = tk.Label(frm_right_bottom_right, textvariable=varmdotWater, bg='white',
    #                           font=('Arial', 12), width=15, height=2)
    # labelmdotWater.pack()

    # on_click_loop = False

    # def btn_cmd_loop(func):
    #     global on_click_loop
    #     if on_click_loop == False:
    #         on_click_loop = True
    #         varScan.set('start2scan')
    #         func()
    #     else:
    #         on_click_loop = False
    #         varScan.set('stop2scan')


#     def btn_cmd_one(func):
#         func()

#     def good():
#         mdotWater = varmdotWater.get()
# #        print(mdotWater, type(mdotWater))
#         SM_dia.update_mdotWater(float(mdotWater))
#         data.update_mdotWater(float(mdotWater))
#         labelmdotWater.config(text=str(mdotWater))

    # buttonScan = tk.Button(
    #     frm_right_bottom_left,
    #     text='click me',
    #     width=15, height=2,
    #     command=lambda: btn_cmd_loop(update_diagram))
    # buttonScan.pack()

#     g = tk.Radiobutton(frm_right_bottom_right, text='熱水大流量',  variable=varmdotWater, value=0.29, \
#                   command=good)
#     g.pack()
#     gg = tk.Radiobutton(frm_right_bottom_right, text='熱水中流量', variable=varmdotWater, value=0.23, \
#                   command=good)
#     gg.pack()
#     ggg = tk.Radiobutton(frm_right_bottom_right, text='熱水小流量', variable=varmdotWater, value=0.17, \
#                   command=good)
#     ggg.pack()

    window.mainloop()
